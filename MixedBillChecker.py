from fileinput import filename
from http.client import FOUND
from lib2to3.pgen2.token import NEWLINE
from msilib.schema import Icon
import openpyxl
import datetime
import os
import pyperclip    # for copying and pasting the PDF to notepad
import keyboard     # automate keyboard input
import time         # gives time between automated keyboard actions
import shutil
import warnings






# access the folder that contains the bills you want to extract dominion gas bills from
dir_list = os.listdir('K:\\meters\\Utility Bills\\Current Month\\1_Received\\Patrick\\MixedTest')


# identifies and puts all dominion gas bills into file_list as a list
file_list = []                       # contains strings of the file names of each file in the directory (aka folder)
#len(dir_list)
for i in range(0, len(dir_list)):
    if (('E' in dir_list[i]) and (len(dir_list[i]) > 19) or ('W' in dir_list[i]) and (len(dir_list[i]) > 19)) and ('.pdf' in dir_list[i]):      # len(dir_list[i]) > 19 makes it so it doesn't grab PDFs with less than one bill in them
        file_list.append(dir_list[i])


# opens the PDF and copies the text into a list, then transforms each element into a list, separated by newlines
content_list = []    # each element will contain a list of each bill's contents, separated by newlines.
                                    # this makes it so we can access relevant information by index in each bill

# open each PDF and copy the contents into a list one by one
for i in range(0, len(file_list)):
    #for j in range(0, len(file_list[i])):
    os.startfile('K:\\meters\\Utility Bills\\Current Month\\1_Received\\Patrick\\MixedTest\\' + file_list[i])
    time.sleep(1)
    keyboard.press_and_release('ctrl+a')
    time.sleep(1)
    keyboard.press_and_release('ctrl+c')
    time.sleep(1)
    os.system('taskkill /im acrobat.exe /f')        # closes Adobe Acrobat
    content = pyperclip.paste()                     # pastes the copied PDF contents into a string
    bill = content.splitlines()                     # transforms the string into a list, separated by newlines
    content_list.append(bill)           # appends the list to the list of bills

print('Bills being checked:')
print(file_list, '\n')
print()


print("Checking has finished. Here are the values that did not match up: \n")


# pdf variables
# global meternum_pdf
# global beg_read_pdf
# global end_read_pdf
# global dth_pdf
# global vol_mul_pdf
# global cost_pdf
# global start_date_pdf
# global end_date_pdf

def IterateWatExcel(sheet):
    global wat_meternum_xl, wat_start_date_xl, wat_end_date_xl, wat_beg_read_xl, wat_end_read_xl, wat_cost_xl, wat_kgal_xl, found

    # wat_meternum_xl[0] == 'Not assigned' and wat_meternum_xl[1] == 'meternum'
    wat_meternum_xl      = ['Not assigned', 'meternum']
    wat_start_date_xl    = ['Not assigned', 'start_date']
    wat_end_date_xl      = ['Not assigned', 'end_date']
    wat_beg_read_xl      = ['Not assigned', 'beg_read']
    wat_end_read_xl      = ['Not assigned', 'end_read']
    wat_cost_xl          = ['Not assigned', 'cost']
    wat_kgal_xl          = ['Not assigned', 'kgal']

    for row in sheet.iter_rows(min_row = 2, max_row = 2, min_col = sheet.min_column, max_col = sheet.max_column):
        for cell in row:        # note: the variable "cell" is the cell address
            if str(cell.value) == str(wat_meternum_pdf):
                found = True
                wat_meternum_xl = str(cell.value)
                wat_start_date_xl = str(sheet.cell(row = cell.row + 5, column = cell.column - 4).value)
                wat_end_date_xl = str(sheet.cell(row = cell.row + 4, column = cell.column - 4).value)
                wat_beg_read_xl = str(sheet.cell(row = cell.row + 5, column = cell.column - 3).value)
                wat_end_read_xl = str(sheet.cell(row = cell.row + 4, column = cell.column - 3).value)
                wat_kgal_xl = str(sheet.cell(row = cell.row + 4, column = cell.column - 2).value)
                wat_cost_xl = str(sheet.cell(row = cell.row + 4, column = cell.column + 1).value)
                break
            else:
                continue


# xl[1] is the name of the check ("cost check" or "beg read check")
def Print_Checks(pdf, xl):
    if pdf == xl[0]:
        # print(xl[1] + " succeeded -- PDF: " + pdf + " -- Excel: " + xl)
        return True
    else:
        print(xl[1] + " failed -- PDF: " + pdf + " -- Excel: " + xl)
        return False


# checks the PDF values to the Excel values
def Check_Values(bill, meternum_pdf, start_date_pdf, end_date_pdf, cost_pdf, start_read_pdf, end_read_pdf, check1_pdf, check2_pdf,
                       meternum_xl , start_date_xl , end_date_xl , cost_xl , start_read_xl , end_read_xl , check1_xl , check2_xl):
    global matchingBills
    global nonmatchingBills

    #check_failure_list = []                  # this is a list that will contain bills that contained one or more failed checks
    
    # locate, return, and compare START READ in excel to PDF
    meternum_check = Print_Checks(meternum_pdf, meternum_xl)
    
    # locate, return, and compare START DATE in excel to PDF
    start_date_check = Print_Checks(start_date_pdf, start_date_xl)

    # locate, return, and compare END DATE in excel to PDF
    end_date_check = Print_Checks(end_date_pdf, end_date_xl)

    # locate, return, and compare COST in excel to PDF
    cost_check = Print_Checks(cost_pdf, cost_xl)

    # locate, return, and compare START READ in excel to PDF
    start_read_check = Print_Checks(start_read_pdf, start_read_xl)

    # locate, return, and compare END READ in excel to PDF
    end_read_check = Print_Checks(end_read_pdf, end_read_xl)

    check1_check = Print_Checks(check1_pdf, check1_xl)

    check2_check = Print_Checks(check2_pdf, check2_xl)

    # confirm all checks came out as true
    if (start_date_check == True and end_date_check == True and cost_check == True and meternum_check == True and start_read_check == True and end_read_check == True and check1_check == True and check2_check == True):
        matchingBills.append(bill)
    else:
        nonmatchingBills.append(bill)


# this is used to search for specific text in the bill
# the text will be located at content_list[iCount][jCount + x]
def FindText(text):
    x = 0
    while text not in content_list[iCount][jCount + x]:
        x = x + 1
        if text in content_list[iCount][jCount + x]:
            return x


# formats the cost from the PDF to match the format found in Excel
def FormatCostPDF(cost_pdf):
    if cost_pdf[-1] == '0':
        cost_pdf = cost_pdf[:-1]
    if cost_pdf[-1] == '0':
        cost_pdf = cost_pdf[:-1]
    if cost_pdf[-1] == '.':
        cost_pdf = cost_pdf[:-1]
    cost_pdf = cost_pdf.replace('$', '')
    return cost_pdf


def DeleteCommas(value):
    if ',' in value:
        value.replace(',', '')
    return value









# checking begins
for iCount in range(0, len(content_list)):
    bill_count = 0
    # print the number of bills in this PDF
    print('In PDF: ' + file_list[iCount])
    numBills = file_list[iCount].count('E') + file_list[iCount].count('W') + file_list[iCount].count('R') + file_list[iCount].count('S')

    # remove the date and ".pdf" from the bills, then split into a list of individual bill names
    bill_list = file_list[iCount][10:]
    bill_list = bill_list[:-4]
    bill_list = bill_list.split('_')
    print(bill_list)

    matchingBills = []
    nonmatchingBills = []
    
    for jCount in range(0, len(content_list[iCount])):

        # STORM DRAIN BILL
        '''
        # cost
        if content_list[iCount][jCount] == 'Storm Water Charge':
            if '$' in content_list[iCount][jCount - 1]:
                std_cost_pdf = content_list[iCount][jCount - 1].split('$')[1]
                std_cost_pdf = FormatCostPDF(std_cost_pdf)
                print("std_cost_pdf: " + std_cost_pdf)

            # get the bill number from pdf
            for x_count in range(0, len(bill_list)):
                if 'R' in bill_list[x_count]:
                    storm_bill_pdf = bill_list[x_count]
                    # the PDF names 37R and 38R '85R' and 86R', so this renames them as found in the xl
                    if storm_bill_pdf == '37R':
                        storm_bill_pdf = '85R'
                    if storm_bill_pdf == '38R':
                        storm_bill_pdf = '86R'
                    print("storm bill: " + storm_bill_pdf)
        
            from openpyxl import load_workbook
            # go here for an explanation of this warning code stuff: https://stackoverflow.com/questions/34322231/python-2-7-openpyxl-userwarning
            warnings.simplefilter("ignore")
            wb = load_workbook(filename = 'K:\\meters\\readings\\std.xlsm', data_only = True)
            ws_1 = wb['StormDrain1']
            warnings.resetwarnings()

            # iterate through Row 4 and look for the meter nickname (i.e. 17R) (there is no meter number) in the std excel document
            found = False
            for row in ws_1.iter_rows(min_row = 4, max_row = 4, min_col = ws_1.min_column, max_col = ws_1.max_column):
                for cell in row:        # note: the variable "cell" is the cell address
                    if str(cell.value) == str(storm_bill_pdf):
                        found = True
                        storm_bill_xl = cell.value
                        std_cost_xl = ws_1.cell(row = cell.row + 2, column = cell.column + 3).value
                        print("std_cost_xl: " + str(std_cost_xl))
                        break
                    else:
                        continue
            if str(std_cost_pdf) == str(std_cost_xl):
                matchingBills = '' + storm_bill_pdf
            
            if str(std_cost_pdf) != str(std_cost_xl):
                nonmatchingBills = '' + str(storm_bill_pdf)

        '''
        # END OF STORM DRAIN










        # WATER BILL
        if "1 TGAL = 1000 Gallons" in content_list[iCount][jCount]:
            print('Water:')
            for x in range(0, len(bill_list)):
                if 'W' in bill_list[x]:
                    wat_bill_pdf = bill_list[x]


            # METER NUMBER
            x = 0
            while 'Usage for meter: ' not in content_list[iCount][jCount + x]:
                x = x + 1
            if "Usage for meter: " in content_list[iCount][jCount + x]:
                wat_meternum_pdf = str(content_list[iCount][jCount + x].split('Usage for meter: ')[1])


            # START/END DATE
            # will be located several indices below "1 TGAL = 1000 Gallons"
            x = 0
            while content_list[iCount][jCount - x].count('/') != 4:      # keep looking until we find 4 '/' characters in the line; this will have the start and end date
                x = x + 1
            if 'to' in content_list[iCount][jCount - x]:
                dates = str((content_list[iCount][jCount - x])).split(' to ')
                wat_start_date_pdf = dates[0]
                wat_end_date_pdf = dates[1]


            # START READ
            # there are two instances of "Usage for meter: " in the water bill; the start read is 2 indexes after the 2nd instance of "Usage for meter: "
            x = 0
            y = 0
            while y < 2:
                x = x + 1
                if "Usage for meter: " in content_list[iCount][jCount + x]:
                    y = y + 1
            wat_start_read_pdf = str(content_list[iCount][jCount + x + 2])
            wat_start_date_pdf = DeleteCommas(wat_start_date_pdf)
            # print("start read:")


            # END READ
            if bill_list.count('W') == 1:   # the process for a bill with only one water meter
                x = 0
                while "????????????????????" not in content_list[iCount][jCount - x]:
                    x = x + 1
                if "????????????????????" in content_list[iCount][jCount - x]:
                    # the end read will either be 4 indexes lower than the text "Commercial Water Service" or "Industrial Water Service" (whichever one this bill has), OR there will be some stuff in the way that we need to get rid of
                    x = 1
                    print('end read:')
                    while 'Municipal Use Tax' in content_list[iCount][jCount - x] or 'Base Charge' in content_list[iCount][jCount - x] or 'Block' in content_list[iCount][jCount - x]:
                        x = x + 1
                    wat_end_read_pdf = str(content_list[iCount][jCount - x])
                    # the xl doesn't have commas so delete the PDF's comma
                    wat_end_read_pdf = DeleteCommas(wat_end_read_pdf)
            
            if bill_list.count("W") > 1:    # the process for a bill with more than one water meter
                # the end read will be 1 index lower than the date for this bill
                while content_list[iCount][jCount - x].count('/') != 4:      # keep looking until we find 4 '/' characters in the line; this will have the start and end date
                    x = x + 1
                wat_end_read_pdf = str(content_list[iCount][jCount - x - 1])
                wat_end_read_pdf = DeleteCommas(wat_end_read_pdf)


            # KGAL (USAGE)
            # the kgal (usage) shows up 4 indices higher than the text "Year" (there are several "Year"s; this is the first one after "1 TGAL = 1000 Gallons")
            x = 0
            while "Year" not in content_list[iCount][jCount + x]:
                x = x + 1
            if "Year" in content_list[iCount][jCount + x]:
                wat_kgal_pdf = str(content_list[iCount][jCount + x + 4])


            # COST
            # grabs the cost at the next time a cost is listed
            x = 0
            while "$" not in content_list[iCount][jCount + x]:
                x = x + 1
            if "$" in content_list[iCount][jCount + x]:
                wat_cost_pdf = str(content_list[iCount][jCount + x])
                wat_cost_pdf = FormatCostPDF(wat_cost_pdf)


            

            # find matching values from XL
            from openpyxl import load_workbook
            # go here for an explanation of this warning code stuff: https://stackoverflow.com/questions/34322231/python-2-7-openpyxl-userwarning
            warnings.simplefilter("ignore")
            wb = load_workbook(filename = 'K:\\meters\\readings\\Patrick\\wat.xlsx', data_only = True)
            ws_1 = wb['WATER#1.XLS']
            ws_2 = wb['WATER#2.XLS']
            warnings.resetwarnings()

            # iterate through Row 2 and look for the meter number in the std excel document
            found = False
            IterateWatExcel(ws_1)
            if found == False:
                IterateWatExcel(ws_2)
            

            Check_Values(wat_bill_pdf, wat_meternum_pdf, wat_start_date_pdf, wat_end_date_pdf, wat_cost_pdf, wat_start_read_pdf, wat_end_read_pdf, wat_kgal_pdf, "None",
                                       wat_meternum_xl
            )


        # END OF WATER BILL






    print('These bills match:')
    print(matchingBills)
    print('These bills do not match:')
    print(nonmatchingBills)
    print()
        
        


















# # locate, return, and compare START DATE in excel to PDF
#     #check_failure_list = []                  # this is a list that will contain 
#     if str(start_date_pdf) == str(start_date_xl):
#         start_date_check = True
#         # print("start date check succeeded -- PDF: " + start_date_pdf + " -- Excel: " + start_date_xl)
#     else:
#         start_date_check = False
#         #check_failure_list.append(str("Bill " + provo_water_file_list[iCount] + ": start date check failed"))
#         #print("start date check failed -- PDF: " + str(start_date_pdf) + " -- Excel: " + str(start_date_xl))
#         print('start date check failed: ' + 'PDF: ' + start_date_pdf + 'Excel: ' + start_date_xl)

    
#     # locate, return, and compare END DATE in excel to PDF
#     if end_date_pdf == wat_end_date_xl:
#         end_date_check = True
#         #print("end date check succeeded -- PDF: " + end_date_pdf + " -- Excel: " + end_date_xl)
#     else:
#         end_date_check = False
#         print('end date check failed: ' + 'PDF: ' + end_date_pdf + 'Excel: ' + end_date_xl)

#     # locate, return, and compare COST in excel to PDF
#     if cost_pdf == cost_xl:
#         cost_check = True
#         #print("cost check succeeded -- PDF: " + cost_pdf + " -- Excel: " + cost_xl)
#     else:
#         cost_check = False
#         #print("cost check failed -- PDF: " + cost_pdf + " -- Excel: " + cost_xl)
#         print('cost check failed: ' + 'PDF: ' + cost_pdf + 'Excel: ' + cost_xl)


#     # locate, return, and compare START READ in excel to PDF
#     if meternum_pdf == meternum_xl:
#         meternum = True
#         #print("meternum succeeded -- PDF: " + meternum_pdf + " -- Excel: " + meternum_xl)
#     else:
#         meternum = False
#         #check_failure_list.append(str("Bill " + provo_water_file_list[iCount] + ": beginning read check failed"))
#         print('meternum failed: ' + 'PDF: ' + meternum_pdf + 'Excel: ' + meternum_xl)
        

#     # locate, return, and compare START READ in excel to PDF
#     if check5_pdf == check5_xl:
#         check5 = True
#         #print("meternum succeeded -- PDF: " + check5_pdf + " -- Excel: " + check5_xl)
#     else:
#         check5 = False
#         #check_failure_list.append(str("Bill " + provo_water_file_list[iCount] + ": beginning read check failed"))
#         print('check5 failed: ' + 'PDF: ' + check5_pdf + 'Excel: ' + check5_xl)


#     # locate, return, and compare START READ in excel to PDF
#     if check6_pdf == check6_xl:
#         check6 = True
#         #print("meternum succeeded -- PDF: " + check6_pdf + " -- Excel: " + check6_xl)
#     else:
#         check6 = False
#         #check_failure_list.append(str("Bill " + provo_water_file_list[iCount] + ": beginning read check failed"))
#         print('check6 failed: ' + 'PDF: ' + check6_pdf + 'Excel: ' + check6_xl)